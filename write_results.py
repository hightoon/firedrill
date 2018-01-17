import os.path
import xlsxwriter
from openpyxl import Workbook, load_workbook

import sys
reload(sys)
sys.setdefaultencoding("utf-8")

def write(fields, results, sn, expected=None, stats=None, ttboss=None, 
          uniqtops=None, external=None, erroregs=None, total_valid_regs=0):
    wb = xlsxwriter.Workbook('fd_mngm_tr.xlsx')
    wb.encoding = 'utf-8'
    bold = wb.add_format({'bold': 1})
    ws = wb.add_worksheet(sn)
    redbold = wb.add_format({'bold': 1, 'font_color': 'red'})
    warn = wb.add_format({'bold': 1, 'font_color': 'orange'})
    greenbold = wb.add_format({'bold': 1, 'font_color': 'green'})
    bluebold = wb.add_format({'bold': 1, 'font_color': 'blue'})
    row = 0 
    for col, f in enumerate(fields):
        ws.write(row, col, f, bold)
    row += 1
    mgr1, mgr2, mgr3, mgr4 = None, None, None, None
    pre1, pre2, pre3, pre4 = None, None, None, None
    lowest = 5
    last = None
    for result in results:
        mgr4, mgr3, mgr2, mgr1 = result['manager-4'], result['manager-3'], \
                                 result['manager-2'], result['manager-1']
        if pre4 is None:
            pre4 = mgr4
        if mgr4 != pre4:
            if len(last['managers']) > 4:
                try:
                    stat = stats[pre4]
                except KeyError:
                    stat = 0
                ws.write('D%d'%(row+1,), pre4, bold)
                ws.write('E%d'%(row+1,), 'Expected No.: %d'%(expected[pre4],), bold)
                fmt = bold
                if stat < expected[pre4]:
                    fmt = warn
                ws.write('F%d'%(row+1,), 'Actual No.: %d'%(stat,), fmt)
                ws.set_row(row, None, None, {'level': 4, 'hidden': True, 'collapsed': True})
                row += 1
            pre4 = mgr4
        if pre3 is None:
            pre3 = mgr3
        if mgr3 != pre3:
            if len(last['managers']) > 3:
                try:
                    stat = stats[pre3]
                except KeyError:
                    stat = 0
                ws.write('C%d'%(row+1,), pre3, bold)
                ws.write('D%d'%(row+1,), 'Expected No.: %d'%(expected[pre3],), bold)
                fmt = bold
                if stat < expected[pre3]:
                    fmt = warn
                ws.write('E%d'%(row+1,), 'Actual No.: %d'%(stat,), fmt)
                ws.set_row(row, None, None, {'level': 3, 'hidden': True, 'collapsed': True})
                row += 1
            pre3 = mgr3
        if pre2 is None:
            pre2 = mgr2
        if mgr2 != pre2:
            if len(last['managers']) > 2:
                try:
                    stat = stats[pre2]
                except KeyError:
                    stat = 0
                ws.write('B%d'%(row+1,), pre2, bold)
                ws.write('C%d'%(row+1,), 'Expected No.: %d'%(expected[pre2],), bold)
                fmt = bold
                if stat < expected[pre2]:
                    fmt = warn
                ws.write('D%d'%(row+1,), 'Actual No.: %d'%(stat,), fmt)
                ws.set_row(row, None, None, {'level': 2, 'hidden': True, 'collapsed': True})
                row += 1
            pre2 = mgr2
        if pre1 is None:
            pre1 = mgr1
        if mgr1 != pre1:
            try:
                stat = stats[pre1]
            except KeyError:
                stat = 0
            if pre1 in ttboss:
                ws.write('A%d'%(row+1,), pre1, bluebold)
            else:
                ws.write('A%d'%(row+1,), pre1, bold)
            ws.write('B%d'%(row+1,), 'Expected No.: %d'%(expected[pre1],), bold)
            fmt = bold
            if stat < expected[pre1]:
                fmt = warn
            ws.write('C%d'%(row+1,), 'Actual No.: %d'%(stat,), fmt)
            ws.set_row(row, None, None, {'level': 1, 'hidden': True, 'collapsed': True})
            row += 1
            pre1 = mgr1
        for col, f in enumerate(fields):
            ws.write(row, col, result[f])
            if len(result['managers']) > 4:
                ws.set_row(row, None, None, {'level': lowest, 'hidden': True, 'collapsed': True})
            else: # for level-2, 3, 4, set it as it is
                ws.set_row(row, None, None, {'level': len(result['managers']), 'hidden': True, 'collapsed': True})
        row += 1
        last = result
    try:
        stat = stats[pre1]
    except KeyError:
        stat = 0
    ws.write('A%d'%(row+1,), pre1, bold)
    ws.write('B%d'%(row+1,), 'Expected No.: %d'%(expected[pre1],), bold)
    ws.write('C%d'%(row+1,), 'Actual No.: %d'%(stat,), bold)
    ws.set_row(row, None, None, {'level': 1, 'hidden': True, 'collapsed': True})
    row += 1
    for v in external:
        ws.write('A%d'%(row+1,), v.person.company)
        ws.write('B%d'%(row+1,), v.nickname)
        ws.set_row(row, None, None, {'level': 2, 'hidden': True, 'collapsed': True})
        row += 1
    ws.write('A%d'%(row+1,), 'Total Non-Employee Num.: %d'%(len(external),))
    ws.set_row(row, None, None, {'level': 1, 'hidden': True, 'collapsed': True})
    row += 1
    ws.write('A%d'%(row+1,), 'Expected Total: %d'%(len(results)), redbold)
    ws.write('B%d'%(row+1,), 'Total Valid Registrations: %d'%(total_valid_regs,), redbold)
    ws.set_row(row, None, None, {'collapsed': True})
    row += 1
    ws.write('A%d'%(row+1,), 'manager-1 in blue = she/he is not in TT but her/his manager in TT')

    if erroregs is not None:
        err_ws = wb.add_worksheet('abnormal registration')
        err_row = 0
        err_fields = ['employment type', 'registration info', 'wechat nickname']
        for col, f in enumerate(err_fields):
            err_ws.write(err_row, col, f, bold)
        err_row += 1
        reginfos = []
        for reg in erroregs:
            if reg.employment == 'nsb':
                reginfos.append([reg.employment, reg.person.uidnum, reg.nickname])
            else:
                reginfos.append([reg.employment, reg.person.company, reg.nickname])
        for reginfo in reginfos:
            for col, v in enumerate(reginfo):
                err_ws.write(err_row, col, v)
            err_row += 1
    wb.close()


def write_(fields, ldap_q_results, sn):
    '''xlsfile = 'firedrill.xlsx'
    if not os.path.exists(xlsfile):
        wb = Workbook(write_only=True)
    else:
        wb = load_workbook(filename = xlsfile)
    try:
        ws = wb[sn]
    except:
        ws = wb.create_sheet(sn)
    
    ws.append(fields)
    for result in ldap_q_results:
        ws.append([result[k] for k in fields])
    wb.save(xlsfile)'''
    if sn == 'management tree sheet':
        tm_wb = xlsxwriter.Workbook('fd_mngm_tr.xlsx')
        ws = tm_wb.add_worksheet('management tree')
        row = 0
        for col, f in enumerate(fields):
            ws.write(row, col, f, bold)
        row += 1
        topmgr = ''
        topsub = 0
        mgr2 = ''
        sub2 = 0
        mgr3 = ''
        sub3 = 0
        mgr4 = ''
        sub4 = 0
        pretop = ''
        pre2 = ''
        pre3 = None
        pre4 = None
        levels = 0
        prelevels = 0
        bold = tm_wb.add_format({'bold': 1})
        for d in ldap_q_results:
            topmgr = d['manager-1']
            mgr2 = d['manager-2']
            mgr3 = d['manager-3']
            mgr4 = d['manager-4']
            levels = len(d['managers'])
            cur_level = 5
            if pretop == '':
                pretop = topmgr
            if pre2 == '':
                pre2 = mgr2
            if pre3 is None:
                pre3 = mgr3
            if pre4 is None:
                pre4 = mgr4
            if pre4 != mgr4:
                is_tail = levels == 4
                if prelevels > 4:
                    ws.write('D%d'%(row+1,), pre4, bold)
                    ws.write('E%d'%(row+1,), 'Expected No.: ', bold)
                    ws.write('F%d'%(row+1,), 'Actual No.: ', bold)
                    ws.set_row(row, None, None, {'level': 4, 'hidden': True, 'collapsed': True})
                    row += 1
                elif is_tail:
                    cur_level = 4
                sub4 = 0
            else:
                sub4 += 1
            if pre3 != mgr3:
                is_tail = levels == 3
                if prelevels > 3:
                    ws.write('C%d'%(row+1,), pre3, bold)
                    ws.write('D%d'%(row+1,), 'Expected No.: ', bold)
                    ws.write('E%d'%(row+1,), 'Actual No.: ', bold)
                    ws.set_row(row, None, None, {'level': 3, 'hidden': True, 'collapsed': True})
                    row += 1
                elif is_tail:
                    #ws.set_row(row, None, None, {'level': 3})
                    cur_level = 3
                sub3 = 0 # clear sub counter
            else:
                sub3 += 1
            if pre2 != mgr2:
                is_tail = levels == 2 and sub2 == 1
                if levels > 2:
                    ws.write('B%d'%(row+1,), pre2, bold)
                    ws.write('C%d'%(row+1,), 'Expected No.: ', bold)
                    ws.write('D%d'%(row+1,), 'Actual No.: ', bold)
                    ws.set_row(row, None, None, {'level': 2, 'hidden': True, 'collapsed': True})
                    row += 1
                elif is_tail:
                    #ws.set_row(row, None, None, {'level': 2})
                    cur_level = 2
                sub2 = 0
            else:
                sub2 += 1
            if pretop != topmgr:
                ws.write('A%d'%(row+1,), pretop, bold)
                ws.write('B%d'%(row+1,), 'Expected No.: ', bold)
                ws.write('C%d'%(row+1,), 'Actual No.: ', bold)
                ws.set_row(row, None, None, {'level': 1, 'hidden': True, 'collapsed': True})
                row += 1
            else:
                topsub += 1
            for col, f in enumerate(fields):
                ws.write(row, col, d[f])
                if len(d['managers']) == 2:
                    l = 2
                else:
                    l = 5
                ws.set_row(row, None, None, {'level': l, 'hidden': True})
            levels = len(d['managers'])
            pretop = topmgr
            pre2 = mgr2
            pre3 = mgr3
            pre4 = mgr4
            if sub2 == 0:
                sub2 += 1
            if sub3 == 0:
                sub3 += 1
            if sub4 == 0:
                sub4 += 1
            row += 1
        row += 1
        ws.set_row(row, None, None, {'collapsed': True})
        tm_wb.close()
        

def group(xlsfile, sheet, row, level):
    wb = xlsxwriter.Workbook(xlsfile)
    ws = wb.load_worksheet(sheet)


if __name__ == '__main__':
    # test code
    fields = ['uid', 'uidNumber']
    results = [{'uid': 'haitchen', 'uidNumber': '61418924'},
               {'uid': 'haizzhang', 'uidNumber': '61418024'}]
    write(fields, results, 'testsheet') 
